from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo


def build_excel_report(
    agency_name,
    agency_website,
    agency_email,
    client_name,
    client_website,
    score,
    score_label,
    results,
    pages,
    broken_links,
    technical_files,
):

    wb = Workbook()

    # =====================================================
    # THEME
    # =====================================================

    navy = "172554"
    blue = "2563EB"
    light_blue = "EFF6FF"
    green = "DCFCE7"
    green_text = "166534"
    orange = "FEF3C7"
    orange_text = "92400E"
    red = "FEE2E2"
    red_text = "991B1B"
    gray = "F8FAFC"
    border_color = "E2E8F0"
    dark = "111827"
    white = "FFFFFF"

    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    # =====================================================
    # HELPER FUNCTIONS
    # =====================================================

    def style_title(ws, title, subtitle=None):

        ws.merge_cells("A1:F1")

        cell = ws["A1"]
        cell.value = title
        cell.font = Font(
            size=20,
            bold=True,
            color=white,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=navy,
        )
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

        ws.row_dimensions[1].height = 34

        if subtitle:

            ws.merge_cells("A2:F2")

            sub = ws["A2"]
            sub.value = subtitle
            sub.font = Font(
                size=10,
                color="64748B",
            )

    def style_headers(ws, row, columns):

        for col in range(1, columns + 1):

            cell = ws.cell(
                row=row,
                column=col,
            )

            cell.font = Font(
                bold=True,
                color=white,
            )

            cell.fill = PatternFill(
                "solid",
                fgColor=blue,
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

            cell.border = thin_border

        ws.row_dimensions[row].height = 28

    def auto_width(ws):

        for column_cells in ws.columns:

            length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                try:

                    value_length = len(
                        str(cell.value)
                    )

                    length = max(
                        length,
                        value_length
                    )

                except Exception:
                    pass

            ws.column_dimensions[
                column_letter
            ].width = min(
                max(length + 3, 12),
                55
            )

    def add_table(ws, start_row, end_row, end_col, name):

        if end_row <= start_row:
            return

        ref = (
            f"A{start_row}:"
            f"{get_column_letter(end_col)}"
            f"{end_row}"
        )

        table = Table(
            displayName=name,
            ref=ref,
        )

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        table.tableStyleInfo = style

        ws.add_table(table)

    # =====================================================
    # EXECUTIVE SUMMARY
    # =====================================================

    ws = wb.active
    ws.title = "Executive Summary"

    style_title(
        ws,
        "Agency SEO Audit Report",
        "Professional client-ready SEO performance report",
    )

    summary_data = [
        ("Agency", agency_name or "—"),
        ("Agency Website", agency_website or "—"),
        ("Agency Email", agency_email or "—"),
        ("Client", client_name or "—"),
        ("Client Website", client_website or "—"),
        ("Overall SEO Score", f"{score}/100"),
        ("SEO Status", score_label or "—"),
        ("Total Checks", len(results)),
        (
            "Critical Issues",
            sum(
                1
                for r in results
                if r.get("severity") == "Critical"
            ),
        ),
        (
            "Warnings",
            sum(
                1
                for r in results
                if r.get("severity") == "Warning"
            ),
        ),
        (
            "Passed Checks",
            sum(
                1
                for r in results
                if r.get("severity") == "Passed"
            ),
        ),
        ("Pages Crawled", len(pages)),
        ("Broken Links", len(broken_links)),
    ]

    row = 4

    for label, value in summary_data:

        ws.cell(
            row=row,
            column=1,
            value=label,
        )

        ws.cell(
            row=row,
            column=2,
            value=value,
        )

        ws.cell(
            row=row,
            column=1,
        ).font = Font(
            bold=True,
            color=dark,
        )

        ws.cell(
            row=row,
            column=1,
        ).fill = PatternFill(
            "solid",
            fgColor=light_blue,
        )

        ws.cell(
            row=row,
            column=1,
        ).border = thin_border

        ws.cell(
            row=row,
            column=2,
        ).border = thin_border

        row += 1

    # Score emphasis

    score_cell = ws["B9"]

    score_cell.font = Font(
        size=18,
        bold=True,
        color=blue,
    )

    # =====================================================
    # FINDINGS SHEET
    # =====================================================

    ws_findings = wb.create_sheet(
        "SEO Findings"
    )

    style_title(
        ws_findings,
        "SEO Findings",
        "Prioritized SEO issues detected during the audit",
    )

    headers = [
        "Severity",
        "Category",
        "Issue",
        "Recommendation",
        "Occurrences",
        "Page / URL",
    ]

    header_row = 4

    for col, header in enumerate(
        headers,
        start=1,
    ):

        ws_findings.cell(
            row=header_row,
            column=col,
            value=header,
        )

    style_headers(
        ws_findings,
        header_row,
        len(headers),
    )

    current_row = 5

    priority = {
        "Critical": 0,
        "Warning": 1,
        "Passed": 2,
    }

    sorted_results = sorted(
        results,
        key=lambda x: priority.get(
            x.get("severity"),
            3,
        ),
    )

    for result in sorted_results:

        severity = result.get(
            "severity",
            "Warning",
        )

        values = [
            severity,
            result.get(
                "category",
                "Other",
            ),
            result.get(
                "issue",
                "",
            ),
            result.get(
                "recommendation",
                "",
            ),
            1,
            (
                result.get("url")
                or result.get("page")
                or result.get("Page")
                or ""
            ),
        ]

        for col, value in enumerate(
            values,
            start=1,
        ):

            cell = ws_findings.cell(
                row=current_row,
                column=col,
                value=value,
            )

            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        severity_cell = ws_findings.cell(
            row=current_row,
            column=1,
        )

        if severity == "Critical":

            severity_cell.fill = PatternFill(
                "solid",
                fgColor=red,
            )

            severity_cell.font = Font(
                bold=True,
                color=red_text,
            )

        elif severity == "Warning":

            severity_cell.fill = PatternFill(
                "solid",
                fgColor=orange,
            )

            severity_cell.font = Font(
                bold=True,
                color=orange_text,
            )

        else:

            severity_cell.fill = PatternFill(
                "solid",
                fgColor=green,
            )

            severity_cell.font = Font(
                bold=True,
                color=green_text,
            )

        current_row += 1

    add_table(
        ws_findings,
        header_row,
        current_row - 1,
        len(headers),
        "SEOFindingTable",
    )

    ws_findings.freeze_panes = "A5"
    ws_findings.auto_filter.ref = (
        f"A4:F{max(current_row - 1, 4)}"
    )

    auto_width(ws_findings)

    # =====================================================
    # PAGES SHEET
    # =====================================================

    ws_pages = wb.create_sheet(
        "Pages"
    )

    style_title(
        ws_pages,
        "Crawled Pages",
        "Pages discovered and analyzed during the SEO audit",
    )

    page_headers = [
        "URL",
        "HTTP Status",
        "Title",
        "Issues",
    ]

    for col, header in enumerate(
        page_headers,
        start=1,
    ):

        ws_pages.cell(
            row=4,
            column=col,
            value=header,
        )

    style_headers(
        ws_pages,
        4,
        len(page_headers),
    )

    row = 5

    for page in pages:

        values = [
            page.get("URL", ""),
            page.get("Status", ""),
            page.get("Title", ""),
            page.get("Issues", 0),
        ]

        for col, value in enumerate(
            values,
            start=1,
        ):

            cell = ws_pages.cell(
                row=row,
                column=col,
                value=value,
            )

            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        row += 1

    add_table(
        ws_pages,
        4,
        row - 1,
        4,
        "PagesTable",
    )

    ws_pages.freeze_panes = "A5"

    auto_width(ws_pages)

    # =====================================================
    # BROKEN LINKS
    # =====================================================

    ws_links = wb.create_sheet(
        "Broken Links"
    )

    style_title(
        ws_links,
        "Broken Links",
        "Broken links discovered during the audit",
    )

    if broken_links:

        keys = set()

        for item in broken_links:
            keys.update(
                item.keys()
            )

        link_headers = list(keys)

        for col, header in enumerate(
            link_headers,
            start=1,
        ):

            ws_links.cell(
                row=4,
                column=col,
                value=header,
            )

        style_headers(
            ws_links,
            4,
            len(link_headers),
        )

        row = 5

        for item in broken_links:

            for col, header in enumerate(
                link_headers,
                start=1,
            ):

                cell = ws_links.cell(
                    row=row,
                    column=col,
                    value=item.get(
                        header,
                        "",
                    ),
                )

                cell.border = thin_border
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

            row += 1

        add_table(
            ws_links,
            4,
            row - 1,
            len(link_headers),
            "BrokenLinksTable",
        )

        ws_links.freeze_panes = "A5"

    else:

        ws_links["A4"] = (
            "No broken links detected."
        )

        ws_links["A4"].font = Font(
            bold=True,
            color=green_text,
        )

        ws_links["A4"].fill = PatternFill(
            "solid",
            fgColor=green,
        )

    auto_width(ws_links)

    # =====================================================
    # TECHNICAL SHEET
    # =====================================================

    ws_technical = wb.create_sheet(
        "Technical SEO"
    )

    style_title(
        ws_technical,
        "Technical SEO",
        "Technical files and website infrastructure checks",
    )

    technical_headers = [
        "Check",
        "Status",
        "HTTP Status",
    ]

    for col, header in enumerate(
        technical_headers,
        start=1,
    ):

        ws_technical.cell(
            row=4,
            column=col,
            value=header,
        )

    style_headers(
        ws_technical,
        4,
        3,
    )

    technical_rows = [
        (
            "robots.txt",
            technical_files.get(
                "robots",
                {},
            ).get(
                "working",
                False,
            ),
            technical_files.get(
                "robots",
                {},
            ).get(
                "status",
                "",
            ),
        ),
        (
            "sitemap.xml",
            technical_files.get(
                "sitemap",
                {},
            ).get(
                "working",
                False,
            ),
            technical_files.get(
                "sitemap",
                {},
            ).get(
                "status",
                "",
            ),
        ),
    ]

    row = 5

    for check, working, status_code in technical_rows:

        status_text = (
            "Found"
            if working
            else "Not Found"
        )

        values = [
            check,
            status_text,
            status_code,
        ]

        for col, value in enumerate(
            values,
            start=1,
        ):

            cell = ws_technical.cell(
                row=row,
                column=col,
                value=value,
            )

            cell.border = thin_border

        status_cell = ws_technical.cell(
            row=row,
            column=2,
        )

        if working:

            status_cell.fill = PatternFill(
                "solid",
                fgColor=green,
            )

            status_cell.font = Font(
                bold=True,
                color=green_text,
            )

        else:

            status_cell.fill = PatternFill(
                "solid",
                fgColor=red,
            )

            status_cell.font = Font(
                bold=True,
                color=red_text,
            )

        row += 1

    add_table(
        ws_technical,
        4,
        row - 1,
        3,
        "TechnicalTable",
    )

    ws_technical.freeze_panes = "A5"

    auto_width(ws_technical)

    # =====================================================
    # GENERAL WORKBOOK SETTINGS
    # =====================================================

    for worksheet in wb.worksheets:

        worksheet.sheet_view.showGridLines = False

        worksheet.sheet_properties.pageSetUpPr.fitToPage = True

        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0

        worksheet.sheet_properties.outlinePr.summaryBelow = True

    # =====================================================
    # RETURN FILE
    # =====================================================

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return output.getvalue()
