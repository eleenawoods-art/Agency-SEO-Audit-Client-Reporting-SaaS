from io import BytesIO
from datetime import datetime
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _rows(value):
    if value is None:
        return []
    if isinstance(value, list):
        return [x for x in value if isinstance(x, dict)]
    if isinstance(value, dict):
        return [value]
    return []


def _safe(value):
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return "" if value is None else str(value)


def _style_sheet(ws):
    ws.freeze_panes = "A2"
    thin = Side(style="thin", color="D9E1F2")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)
    for col in range(1, ws.max_column + 1):
        values = [str(ws.cell(r, col).value or "") for r in range(1, min(ws.max_row, 100) + 1)]
        ws.column_dimensions[get_column_letter(col)].width = min(max(max((len(v) for v in values), default=10) + 2, 12), 55)


def _write_table(wb, title, headers, rows):
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for row in rows:
        ws.append([_safe(x) for x in row])
    _style_sheet(ws)
    return ws


def build_excel_report(agency_name="", agency_website="", agency_email="", client_name="Client", client_website="", score=0, score_label="", results=None, pages=None, broken_links=None, technical_files=None):
    results = results or []
    pages = pages or []
    broken_links = broken_links or []
    technical_files = technical_files or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    summary = [
        ("Agency", agency_name), ("Agency Website", agency_website), ("Agency Email", agency_email),
        ("Client", client_name), ("Website", client_website), ("Report Date", datetime.now().strftime("%d %B %Y")),
        ("SEO Score", score), ("Score Status", score_label),
        ("Critical Issues", sum(1 for x in results if x.get("severity") == "Critical")),
        ("Warnings", sum(1 for x in results if x.get("severity") == "Warning")),
        ("Passed Checks", sum(1 for x in results if x.get("severity") == "Passed")),
        ("Pages Crawled", len(pages)), ("Links Checked", len(broken_links)),
    ]
    ws.append(["Metric", "Value"])
    for row in summary:
        ws.append(list(row))
    _style_sheet(ws)

    _write_table(wb, "Audit Results", ["Severity", "Category", "Issue", "Recommended Fix", "Page / URL"], [
        [x.get("severity"), x.get("category"), x.get("issue") or x.get("message"), x.get("recommendation") or x.get("fix"), x.get("url") or x.get("URL")] for x in results
    ])
    _write_table(wb, "Pages", ["URL", "Status", "Title", "Issues"], [
        [x.get("URL") or x.get("url"), x.get("Status") or x.get("status") or x.get("status_code"), x.get("Title") or x.get("title"), x.get("Issues") or x.get("issues") or 0] for x in pages if isinstance(x, dict)
    ])
    _write_table(wb, "Broken Links", ["Source URL", "URL", "Status", "State", "Redirects", "Final URL", "Error"], [
        [x.get("Source URL") or x.get("source_url"), x.get("URL") or x.get("url"), x.get("Status") or x.get("status"), x.get("State") or x.get("state"), x.get("Redirects") or x.get("redirects"), x.get("Final URL") or x.get("final_url"), x.get("Error") or x.get("error")] for x in broken_links if isinstance(x, dict)
    ])
    tech_rows = []
    for section, value in (technical_files.items() if isinstance(technical_files, dict) else []):
        tech_rows.append([section, _safe(value)])
    _write_table(wb, "Technical", ["Check", "Details"], tech_rows)

    # Put the summary first and make it presentable.
    wb.active = 0
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
